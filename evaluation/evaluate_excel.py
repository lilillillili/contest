import requests
import re
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font

# --- 1. 데이터 준비 단계 함수들 (이전과 동일) ---

def read_links_from_file(filepath):
    """텍스트 파일에서 링크 목록을 읽어옵니다."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        links = [link.strip() for link in content.split(',') if link.strip()]
        print(f"📄 '{os.path.basename(filepath)}'에서 {len(links)}개의 링크를 찾았습니다.")
        return links
    except FileNotFoundError:
        print(f"❌ 오류: '{filepath}' 파일을 찾을 수 없습니다.")
        return []
    return []

def get_filename_from_url(url):
    """구글 드라이브 링크에서 파일명을 추출합니다."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept-Language': 'ko-KR,ko;q=0.9'
    }
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code == 200:
            match = re.search(r'<title>(.*?) - Google Drive</title>', response.text)
            return match.group(1).strip() if match else None
    except requests.RequestException:
        return None

def extract_team_name(filename):
    """파일명에서 팀명을 추출합니다."""
    name_without_ext = os.path.splitext(filename)[0]
    name_cleaned = re.sub(r'^\d+\.\s*', '', name_without_ext)
    team_name = name_cleaned.split('_')[-1] if '_' in name_cleaned else name_cleaned
    return team_name.strip()

# --- 2. 엑셀 파일 업데이트 함수 ---

def update_evaluation_sheet(team_link_map, template_path, output_path):
    """
    기존 엑셀 양식을 읽어 팀명에 맞는 구글 드라이브 하이퍼링크를 추가
    (하이퍼링크 텍스트를 동적으로 변경)
    """
    try:
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active
        print(f"\n📂 '{os.path.basename(template_path)}' 파일을 열었습니다.")
    except FileNotFoundError:
        print(f"❌ 오류: '{template_path}' 파일을 찾을 수 없습니다.")
        return

    hyperlink_font = Font(color="0000FF", underline="single")
    
    matched_count = 0
    # 데이터가 시작되는 8번 행부터 마지막 행까지 순회
    for row in range(8, sheet.max_row + 1):
        team_name_cell = sheet.cell(row=row, column=4)
        team_name = team_name_cell.value

        if team_name and isinstance(team_name, str):
            clean_team_name = team_name.strip()
            link = team_link_map.get(clean_team_name)

            if link:
                link_cell = sheet.cell(row=row, column=6)
                
                # f-string을 사용하여 팀명에 맞는 텍스트를 생성
                link_cell.value = f"{clean_team_name} 심사 파일"
                
                link_cell.hyperlink = link
                link_cell.font = hyperlink_font
                matched_count += 1
                print(f"   ✅ '{clean_team_name}' 팀의 링크를 F{row} 셀에 추가했습니다.")

    print(f"\n🎉 총 {matched_count}개의 링크를 엑셀 파일에 성공적으로 추가했습니다.")

    try:
        workbook.save(output_path)
        print(f"💾 결과가 '{os.path.basename(output_path)}' 파일로 저장되었습니다.")
    except Exception as e:
        print(f"❌ 파일 저장 중 오류 발생: {e}")

# --- 3. 메인 실행 부분 ---

def main():
    print("🎯 심사표 엑셀 자동화 프로그램 v10.0")
    print("=" * 60)

    # --- 설정 영역 ---
    # 🚨 파일들이 있는 폴더 경로를 정확하게 지정
    base_dir = r"파일 경로 정확히 지정"
    
    # 1. 링크가 담긴 텍스트 파일
    link_file_path = os.path.join(base_dir, "linklist.txt")
    
    # 2. 링크를 삽입할 엑셀 양식 파일
    template_excel_path = os.path.join(base_dir, "2025년 대회 예선 심사표.xlsx")
    
    # 3. 최종 결과가 저장될 파일
    output_excel_path = os.path.join(base_dir, "2025년 대회 예선 심사표_결과.xlsx")
    # --- 설정 영역 끝 ---

    # Step 1: linklist.txt에서 링크를 읽어 팀명과 링크의 맵(map)을 생성
    links = read_links_from_file(link_file_path)
    if not links:
        return

    team_link_map = {}
    print("\n🔗 각 링크에 접속하여 파일명과 팀명을 추출합니다...")
    for link in links:
        filename = get_filename_from_url(link)
        if filename:
            team_name = extract_team_name(filename)
            team_link_map[team_name] = link
    print(f"👍 {len(team_link_map)}개의 팀-링크 데이터를 준비했습니다.")

    # Step 2: 준비된 데이터를 바탕으로 엑셀 파일 업데이트
    update_evaluation_sheet(team_link_map, template_excel_path, output_excel_path)

if __name__ == "__main__":
    print("📦 필요 라이브러리: pip install requests pandas openpyxl\n")
    main()